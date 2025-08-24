import os
import pandas as pd
from PyQt6 import QtWidgets, QtCore, QtGui
import pickle # 導入 pickle 模組用於深度複製
import numpy as np

# 檢查是否安裝了統計檢定相關套件
try:
    from scipy import stats
    from scipy.stats import ttest_ind, ttest_rel, f_oneway
    try:
        from statsmodels.stats.multicomp import pairwise_tukeyhsd
        TUKEY_AVAILABLE = True
    except ImportError:
        TUKEY_AVAILABLE = False
        pairwise_tukeyhsd = None
    SCIPY_AVAILABLE = True
except ImportError:
    SCIPY_AVAILABLE = False
    stats = None
    ttest_ind = None
    ttest_rel = None
    f_oneway = None
    TUKEY_AVAILABLE = False
    pairwise_tukeyhsd = None

# 檢查是否安裝了 openpyxl 套件
try:
    import openpyxl
except ImportError:
    openpyxl = None

class ToolMatchingWidget(QtWidgets.QWidget):
    """
    Tool Matching 分析工具：
    - 讀入 CSV 檔案
    - 根據 GroupName + ChartName 分組
    - 根據 characteristic 進行 mean/sigma matching 檢查
    - 顯示不匹配的結果
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        # 設定全域中文字體為微軟正黑體（僅影響本 widget 及其子元件）
        font = QtGui.QFont("Microsoft JhengHei")
        font.setPointSize(10)
        self.setFont(font)
        self.init_ui()


    def init_ui(self):
        self.setWindowTitle("Tool Matching ")
        self.resize(1200, 800)

        # 主佈局
        self.main_layout = QtWidgets.QVBoxLayout(self)
        self.setLayout(self.main_layout)

        # --- 上方控制區 ---
        top_layout_widget = QtWidgets.QWidget()
        top_layout = QtWidgets.QVBoxLayout(top_layout_widget)

        title = QtWidgets.QLabel("<h2 style='color:#34495E;'>Tool Matching Analysis (OSAT) </h2>")
        # 強制套用微軟正黑體於標題（即使是 HTML）
        title_font = QtGui.QFont("Microsoft JhengHei")
        title_font.setPointSize(16)
        title.setFont(title_font)
        top_layout.addWidget(title)

        file_layout = QtWidgets.QHBoxLayout()
        self.file_path_entry = QtWidgets.QLineEdit()
        self.file_path_entry.setPlaceholderText("請選擇一個 CSV 檔案...")
        self.file_path_entry.setReadOnly(True)
        # 加入資料夾符號於「瀏覽檔案...」按鈕
        file_btn = QtWidgets.QPushButton()
        # 使用 emoji 📁 作為 icon，並將文字設為粗體，字體大小與執行按鈕一致
        file_btn.setText("📁 瀏覽檔案...")
        file_btn.setIcon(QtGui.QIcon())  # 移除原本的 QStyle icon
        btn_font = QtGui.QFont("Microsoft JhengHei")
        btn_font.setBold(True)
        btn_font.setPointSize(12)
        file_btn.setFont(btn_font)
        file_btn.setFixedWidth(150)
        file_btn.clicked.connect(self.select_file)
        file_layout.addWidget(self.file_path_entry)
        file_layout.addWidget(file_btn)

        # 新增 temp 按鈕
        temp_btn = QtWidgets.QPushButton("💾 範例")
        temp_btn.setFont(btn_font)
        temp_btn.setFixedWidth(120)
        temp_btn.clicked.connect(self.generate_temp_csv)
        file_layout.addWidget(temp_btn)

        top_layout.addLayout(file_layout)

        # 先插入 mean/sigma index門檻，再插入補滿樣本數欄位
        mean_threshold_layout = QtWidgets.QHBoxLayout()
        self.mean_index_checkbox = QtWidgets.QCheckBox()
        self.mean_index_checkbox.setText("")
        self.mean_index_checkbox.setChecked(False)
        mean_threshold_label = QtWidgets.QLabel("Mean Index 門檻：")
        mean_threshold_label.setFont(QtGui.QFont("Microsoft JhengHei", 11))
        self.mean_index_threshold_spin = QtWidgets.QDoubleSpinBox()
        self.mean_index_threshold_spin.setRange(0, 10)
        self.mean_index_threshold_spin.setValue(1.0)
        self.mean_index_threshold_spin.setSingleStep(0.1)
        self.mean_index_threshold_spin.setFont(QtGui.QFont("Microsoft JhengHei", 11))
        self.mean_index_threshold_spin.setEnabled(False)
        mean_threshold_layout.addWidget(self.mean_index_checkbox)
        mean_threshold_layout.addWidget(mean_threshold_label)
        mean_threshold_layout.addWidget(self.mean_index_threshold_spin)
        mean_threshold_layout.addStretch(1)
        def on_mean_checkbox_changed(state):
            self.mean_index_threshold_spin.setEnabled(self.mean_index_checkbox.isChecked())
        self.mean_index_checkbox.stateChanged.connect(on_mean_checkbox_changed)

        sigma_threshold_layout = QtWidgets.QHBoxLayout()
        self.sigma_index_checkbox = QtWidgets.QCheckBox()
        self.sigma_index_checkbox.setText("")
        self.sigma_index_checkbox.setChecked(False)
        sigma_threshold_label = QtWidgets.QLabel("Sigma Index 門檻：")
        sigma_threshold_label.setFont(QtGui.QFont("Microsoft JhengHei", 11))
        self.sigma_index_threshold_spin = QtWidgets.QDoubleSpinBox()
        self.sigma_index_threshold_spin.setRange(0, 10)
        self.sigma_index_threshold_spin.setValue(2.0)
        self.sigma_index_threshold_spin.setSingleStep(0.1)
        self.sigma_index_threshold_spin.setFont(QtGui.QFont("Microsoft JhengHei", 11))
        self.sigma_index_threshold_spin.setEnabled(False)
        sigma_threshold_layout.addWidget(self.sigma_index_checkbox)
        sigma_threshold_layout.addWidget(sigma_threshold_label)
        sigma_threshold_layout.addWidget(self.sigma_index_threshold_spin)
        sigma_threshold_layout.addStretch(1)
        def on_sigma_checkbox_changed(state):
            self.sigma_index_threshold_spin.setEnabled(self.sigma_index_checkbox.isChecked())
        self.sigma_index_checkbox.stateChanged.connect(on_sigma_checkbox_changed)

        top_layout.addLayout(mean_threshold_layout)
        
        # 新增統計檢定方法選擇
        stats_method_layout = QtWidgets.QHBoxLayout()
        self.use_statistical_test_checkbox = QtWidgets.QCheckBox()
        self.use_statistical_test_checkbox.setText("t-test")
        self.use_statistical_test_checkbox.setChecked(False)
        self.use_statistical_test_checkbox.setFont(QtGui.QFont("Microsoft JhengHei", 11))
        
        self.stats_method_combo = QtWidgets.QComboBox()
        self.stats_method_combo.addItems(["Unpaired t-test", "Paired t-test"])
        self.stats_method_combo.setFixedWidth(220)
        self.stats_method_combo.setFont(QtGui.QFont("Microsoft JhengHei", 10))
        self.stats_method_combo.setEnabled(False)
        
        self.alpha_level_label = QtWidgets.QLabel("顯著水準:")
        self.alpha_level_label.setFont(QtGui.QFont("Microsoft JhengHei", 10))
        self.alpha_level_spin = QtWidgets.QDoubleSpinBox()
        self.alpha_level_spin.setRange(0.001, 0.1)
        self.alpha_level_spin.setValue(0.05)
        self.alpha_level_spin.setSingleStep(0.01)
        self.alpha_level_spin.setDecimals(3)
        self.alpha_level_spin.setFixedWidth(80)
        self.alpha_level_spin.setFont(QtGui.QFont("Microsoft JhengHei", 10))
        self.alpha_level_spin.setEnabled(False)
        
        stats_method_layout.addWidget(self.use_statistical_test_checkbox)
        stats_method_layout.addWidget(self.stats_method_combo)
        stats_method_layout.addWidget(self.alpha_level_label)
        stats_method_layout.addWidget(self.alpha_level_spin)
        stats_method_layout.addStretch(1)
        
        def on_stats_checkbox_changed(state):
            enabled = self.use_statistical_test_checkbox.isChecked()
            self.stats_method_combo.setEnabled(enabled)
            self.alpha_level_spin.setEnabled(enabled)
            # 當使用統計檢定時，禁用 Mean Index
            self.mean_index_checkbox.setEnabled(not enabled)
            if enabled:
                self.mean_index_checkbox.setChecked(False)
                self.mean_index_threshold_spin.setEnabled(False)
        
        self.use_statistical_test_checkbox.stateChanged.connect(on_stats_checkbox_changed)
        top_layout.addLayout(stats_method_layout)
        
        top_layout.addLayout(sigma_threshold_layout)


        fillnum_layout = QtWidgets.QHBoxLayout()
        self.fillnum_checkbox = QtWidgets.QCheckBox()
        self.fillnum_checkbox.setText("")
        self.fillnum_checkbox.setChecked(False)
        fillnum_label = QtWidgets.QLabel("補滿樣本數：")
        fillnum_label.setFont(QtGui.QFont("Microsoft JhengHei", 11))
        self.fillnum_spin = QtWidgets.QSpinBox()
        self.fillnum_spin.setMinimum(1)
        self.fillnum_spin.setMaximum(100)
        self.fillnum_spin.setValue(5)
        self.fillnum_spin.setFont(QtGui.QFont("Microsoft JhengHei", 11))
        self.fillnum_spin.setEnabled(False)
        fillnum_layout.addWidget(self.fillnum_checkbox)
        fillnum_layout.addWidget(fillnum_label)
        fillnum_layout.addWidget(self.fillnum_spin)
        fillnum_layout.addStretch(1)
        def on_fillnum_checkbox_changed(state):
            self.fillnum_spin.setEnabled(self.fillnum_checkbox.isChecked())
        self.fillnum_checkbox.stateChanged.connect(on_fillnum_checkbox_changed)
        top_layout.addLayout(fillnum_layout)

        # 新增資料篩選模式選擇
        filter_layout = QtWidgets.QHBoxLayout()
        self.filter_mode_combo = QtWidgets.QComboBox()
        self.filter_mode_combo.addItems(["全算", "指定日期(一個月mean/半年sigma)", "最新進點(一個月mean/半年sigma)"])
        self.filter_mode_combo.setFixedWidth(260)
        self.filter_mode_combo.setFont(QtGui.QFont("Microsoft JhengHei", 11))
        filter_layout.addWidget(QtWidgets.QLabel("資料篩選模式："))
        filter_layout.addWidget(self.filter_mode_combo)

        self.date_edit = QtWidgets.QDateEdit(QtCore.QDate.currentDate())
        self.date_edit.setDisplayFormat("yyyy-MM-dd")
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setFont(QtGui.QFont("Microsoft JhengHei", 11))
        self.date_edit.setEnabled(False)
        filter_layout.addWidget(QtWidgets.QLabel("指定基準日："))
        filter_layout.addWidget(self.date_edit)
        filter_layout.addStretch(1)
        top_layout.addLayout(filter_layout)

        def on_filter_mode_changed(idx):
            self.date_edit.setEnabled(idx == 1)
        self.filter_mode_combo.currentIndexChanged.connect(on_filter_mode_changed)

        self.run_btn = QtWidgets.QPushButton("🚀 執行分析")
        run_btn_font = QtGui.QFont("Microsoft JhengHei")
        run_btn_font.setBold(True)
        run_btn_font.setPointSize(12)
        self.run_btn.setFont(run_btn_font)
        self.run_btn.clicked.connect(self.run_analysis)
        top_layout.addWidget(self.run_btn)

        # 狀態標籤
        self.status_label = QtWidgets.QLabel("請選擇檔案並點擊開始分析。")
        self.status_label.setFont(QtGui.QFont("Microsoft JhengHei", 10))
        top_layout.addWidget(self.status_label)


        # 可收合的標準說明區塊
        # 1. 摘要說明
        summary_label = QtWidgets.QLabel(
            """
<div style='background-color:#f5f5f5; padding:8px 12px; border-radius:6px; font-size:14px; margin-bottom:4px; font-family:Microsoft JhengHei;'>
  <strong>⚠ 注意：</strong>下表僅顯示異常項目。
  <ul style='margin:8px 0 8px 20px; padding-left:0;'>
    <li><span style='color:#d9534f;'><strong>mean_matching_index ≥ 1</strong></span>：Mean Not Matched (傳統方法)</li>
    <li><span style='color:#d9534f;'><strong>統計檢定顯著</strong></span>：p-value &lt; 0.05 (統計檢定方法)</li>
    <li><span style='color:#d9534f;'><strong>sigma_matching_index ≥ K</strong></span>：Sigma Not Matched</li>
    <li><span style='color:#8a6d3b;'><strong>資料不足</strong></span>：樣本數 &lt; 5, 未進行比較</li>
  </ul>
  <span style='color:#5bc0de;'>可選擇使用統計檢定 (t-test/ANOVA) 取代傳統 Mean Index 方法</span><br>
  <span style='color:#d9534f;'>點擊下方「計算公式」可展開/收合詳細說明。</span>
</div>
            """
        )
        summary_label.setWordWrap(True)
        summary_label.setTextFormat(QtCore.Qt.TextFormat.RichText)
        summary_label.setFont(QtGui.QFont("Microsoft JhengHei", 10))
        top_layout.addWidget(summary_label)

        # 2. 展開/收合按鈕
        self.criterion_toggle_btn = QtWidgets.QToolButton()
        self.criterion_toggle_btn.setText("📘 計算公式 (點擊展開)")
        self.criterion_toggle_btn.setCheckable(True)
        self.criterion_toggle_btn.setChecked(False)
        self.criterion_toggle_btn.setStyleSheet("QToolButton { font-size:13px; color:#344CB7; text-align:left; padding:4px 0; }")
        self.criterion_toggle_btn.setToolButtonStyle(QtCore.Qt.ToolButtonStyle.ToolButtonTextBesideIcon)
        self.criterion_toggle_btn.setArrowType(QtCore.Qt.ArrowType.RightArrow)
        top_layout.addWidget(self.criterion_toggle_btn)

        # 3. 詳細公式內容（預設隱藏，加入滾輪）
        self.criterion_detail_scroll = QtWidgets.QScrollArea()
        self.criterion_detail_scroll.setWidgetResizable(True)
        self.criterion_detail_scroll.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.criterion_detail_scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.criterion_detail_scroll.setMaximumHeight(250)
        # 內容 widget
        detail_content = QtWidgets.QWidget()
        detail_layout = QtWidgets.QVBoxLayout(detail_content)
        detail_layout.setContentsMargins(0, 0, 0, 0)
        detail_label = QtWidgets.QLabel(
            """
<div style="background-color:#f5f5f5; padding:12px; border-radius:6px; font-size:14px; line-height:1.6; font-family:Microsoft JhengHei;">
  <strong>📘 計算公式：</strong>
  <table style="font-size:13px; margin-top:8px; font-family:Microsoft JhengHei;">
    <tr>
      <td style="vertical-align:top; padding-right:8px;"><strong>Mean Matching Index：</strong></td>
      <td>
        <u>兩組比較：</u><br>
        <code>|μ₁ − μ₂| / min(σ₁, σ₂)</code><br>
        <u>多組比較：</u><br>
        <code>|μ − median(μ)| / median(σ)</code>
      </td>
    </tr>
    <tr>
      <td style="vertical-align:top; padding-right:8px;"><strong>Sigma Matching Index：</strong></td>
      <td>
        <u>兩組比較：</u><br>
        <code>σ / min(σ₁, σ₂)</code><br>
        <u>多組比較：</u><br>
        <code>σ / median(σ)</code>
      </td>
    </tr>
    <tr>
      <td style="vertical-align:top; padding-right:8px;"><strong>K 值：</strong></td>
      <td>
        <code>
          n = 樣本數<br>
          n ≤ 4：不比較<br>
          5 ≤ n ≤ 10：K = 1.73<br>
          11 ≤ n ≤ 120：K = 1.414<br>
          n > 120：K = 1.15
        </code>
      </td>
    </tr>
  </table>
  <div style="margin-top:12px; font-size:13px; color:#344CB7;">
    <strong>【篩選模式下的計算說明】</strong><br>
    <ul style="margin:8px 0 8px 20px; padding-left:0;">
      <li>Mean/Std/樣本數：每個 matching_group 取「指定日期往前一個月」內的資料，若不足補到指定筆數（預設5筆，依 UI 可調整），再計算 mean/std/count。</li>
      <li>Median(sigma)：每個 matching_group 取「指定日期往前半年」內的資料，若不足補到指定筆數（預設5筆），每組算 std，再取 median。</li>
      <li>圖表顯示：以「一個月內（補到指定筆數）」的資料為主。</li>
      <li>若同一 group 同一時間有多筆資料，全部納入計算。</li>
    </ul>
    <span style="color:#8a6d3b;">（全算模式則直接用所有資料分組計算，不補點）</span>
  </div>
</div>
            """
        )
        detail_label.setWordWrap(True)
        detail_label.setTextFormat(QtCore.Qt.TextFormat.RichText)
        detail_label.setFont(QtGui.QFont("Microsoft JhengHei", 10))
        detail_layout.addWidget(detail_label)
        self.criterion_detail_scroll.setWidget(detail_content)
        self.criterion_detail_scroll.setVisible(False)
        top_layout.addWidget(self.criterion_detail_scroll)

        # 4. 綁定展開/收合事件（滾輪版）
        def toggle_criterion_detail(checked):
            self.criterion_detail_scroll.setVisible(checked)
            if checked:
                self.criterion_toggle_btn.setText("📘 計算公式 (點擊收合)")
                self.criterion_toggle_btn.setArrowType(QtCore.Qt.ArrowType.DownArrow)
            else:
                self.criterion_toggle_btn.setText("📘 計算公式 (點擊展開)")
                self.criterion_toggle_btn.setArrowType(QtCore.Qt.ArrowType.RightArrow)
        self.criterion_toggle_btn.toggled.connect(toggle_criterion_detail)

        self.main_layout.addWidget(top_layout_widget)

        # --- 結果表格 ---
        self.result_table = QtWidgets.QTableWidget()
        self.result_table.setColumnCount(11) # 調整為 11 列，因為 Need_matching 不在 UI 顯示
        self.result_table.setHorizontalHeaderLabels([
            "GroupName", "ChartName", "Matching Group", "Mean Index", "Sigma Index",
            "K", "Mean", "Sigma", "Mean Median", "Sigma Median", "Sample Size"
        ])
        self.result_table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.result_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.result_table.setAlternatingRowColors(True)
        self.result_table.horizontalHeader().setStretchLastSection(True)
        # 選取整列時有淡藍底且字體顏色為深色，異常欄位紅字不會被蓋掉
        self.result_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #d0d0d0;
            }
            QHeaderView::section {
                background-color: #344CB7;
                color: white;
                padding: 4px;
                font-weight: bold;
            }
            QTableWidget::item {
                background: transparent;
            }
            QTableWidget::item:selected {
                background: #e6f0fa !important;
                color: #222 !important;
            }
        """)
        self.main_layout.addWidget(self.result_table, 1) # 表格佔用更多空間
    def generate_temp_csv(self):
        # 預設範例資料
        data = {
            "GroupName": ["GroupA"],
            "ChartName": ["X"],
            "point_time": ["2023/5/15 14:39"],
            "matching_group": ["A"],
            "point_val": [99.88135943],
            "characteristic": ["Nominal"]
        }
        df = pd.DataFrame(data)
        # 彈出儲存檔案對話框
        save_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "儲存範例 CSV 檔案",
            "tool_matching_input_example.csv",
            "CSV 檔案 (*.csv);;所有檔案 (*.*)"
        )
        if not save_path:
            self.status_label.setText("已取消儲存範例檔案。")
            return
        try:
            df.to_csv(save_path, index=False, encoding="utf-8-sig")
            self.status_label.setText(f"範例 CSV 已儲存到: {save_path}")
        except Exception as e:
            self.status_label.setText(f"範例 CSV 產生失敗: {e}")
    def select_file(self):
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "選擇 CSV 檔案", "", "CSV 檔案 (*.csv);;所有檔案 (*.*)"
        )
        if file_path:
            self.file_path_entry.setText(file_path)
            self.status_label.setText(f"已選擇檔案: {os.path.basename(file_path)}")

    def get_k_value(self, n):
        """根據樣本數量 n 返回 K 值"""
        if n <= 4:  # 樣本數量太少，不進行比較
            return "不比較"  # 返回特殊標記，表示不進行比較
        elif 5 <= n <= 10:
            return 1.73
        elif 11 <= n <= 120:
            return 1.414
        else:
            return 1.15

    def calculate_mean_index(self, mean1, mean2, min_sigma, characteristic):
        """計算 mean matching index，考慮方向性"""
        if min_sigma <= 0:
            return float('inf')
        
        if characteristic == 'up':  # Bigger is better
            return (mean2 - mean1) / min_sigma
        elif characteristic == 'down':  # Smaller is better
            return (mean1 - mean2) / min_sigma
        else:  # Nominal
            return abs(mean1 - mean2) / min_sigma

    def perform_statistical_test(self, data_groups, method="auto", alpha=0.05):
        """
        執行統計檢定
        
        Parameters:
        - data_groups: dict, key 為群組名稱，value 為數據 list
        - method: str, "unpaired", "paired", "auto"
        - alpha: float, 顯著水準
        
        Returns:
        - dict: 包含檢定結果的字典
        """
        if not SCIPY_AVAILABLE:
            return {"error": "SciPy 套件未安裝，無法執行統計檢定"}
        
        group_names = list(data_groups.keys())
        group_data = list(data_groups.values())
        
        # 過濾掉空群組
        valid_groups = [(name, data) for name, data in zip(group_names, group_data) if len(data) >= 2]
        
        if len(valid_groups) < 2:
            return {"error": "有效群組數量不足，無法進行統計檢定"}
        
        valid_names, valid_data = zip(*valid_groups)
        
        result = {
            "method": method,
            "alpha": alpha,
            "groups": valid_names,
            "n_groups": len(valid_groups),
            "significant": False,
            "p_value": None,
            "test_statistic": None,
            "post_hoc": None
        }
        
        try:
            if len(valid_groups) == 2:
                # 兩群比較
                group1_data, group2_data = valid_data[0], valid_data[1]
                
                if method == "paired":
                    # 檢查是否能配對（樣本數相等）
                    if len(group1_data) == len(group2_data):
                        # 配對 t 檢定
                        stat, p_val = ttest_rel(group1_data, group2_data)
                        result["method"] = "Paired t-test"
                    else:
                        return {"error": "配對 t 檢定需要兩組樣本數相等"}
                elif method == "unpaired":
                    # 非配對 t 檢定
                    stat, p_val = ttest_ind(group1_data, group2_data)
                    result["method"] = "Unpaired t-test"
                else:
                    # 預設使用非配對 t 檢定
                    stat, p_val = ttest_ind(group1_data, group2_data)
                    result["method"] = "Unpaired t-test"
                
                result["test_statistic"] = stat
                result["p_value"] = p_val
                result["significant"] = p_val < alpha
                
            else:
                # 三群以上：使用 ANOVA
                stat, p_val = f_oneway(*valid_data)
                result["method"] = "One-way ANOVA"
                result["test_statistic"] = stat
                result["p_value"] = p_val
                result["significant"] = p_val < alpha
                
                # 如果 ANOVA 顯著且有 Tukey HSD 可用，進行事後檢定
                if result["significant"] and TUKEY_AVAILABLE:
                    # 準備 Tukey HSD 的數據格式
                    all_data = []
                    all_groups = []
                    for i, (name, data) in enumerate(valid_groups):
                        all_data.extend(data)
                        all_groups.extend([name] * len(data))
                    
                    try:
                        tukey_result = pairwise_tukeyhsd(all_data, all_groups, alpha=alpha)
                        result["post_hoc"] = {
                            "method": "Tukey HSD",
                            "summary": str(tukey_result),
                            "significant_pairs": []
                        }
                        
                        # 解析顯著的配對
                        for i, row in enumerate(tukey_result.summary().data[1:]):  # 跳過標題行
                            group1, group2, meandiff, p_adj, lower, upper, reject = row
                            if reject:  # reject=True 表示顯著差異
                                result["post_hoc"]["significant_pairs"].append({
                                    "group1": group1,
                                    "group2": group2, 
                                    "p_adj": p_adj,
                                    "mean_diff": meandiff
                                })
                    except Exception as e:
                        result["post_hoc"] = {"error": f"Tukey HSD 失敗: {str(e)}"}
                
        except Exception as e:
            result["error"] = f"統計檢定失敗: {str(e)}"
        
        return result

    def run_analysis(self):
        file_path = self.file_path_entry.text()
        if not file_path or not os.path.exists(file_path):
            self.status_label.setText("請先選擇有效的 CSV 檔案！")
            return

        try:
            df = pd.read_csv(file_path)
        except Exception as e:
            self.status_label.setText(f"讀取檔案失敗: {e}")
            return

        # 檢查必要欄位
        required_cols = ["GroupName", "ChartName", "matching_group", "point_val", "characteristic", "point_time"]
        for col in required_cols:
            if col not in df.columns:
                self.status_label.setText(f"缺少必要欄位: {col}")
                return

        # 轉換 point_time 為 datetime
        try:
            df["point_time"] = pd.to_datetime(df["point_time"])
        except Exception as e:
            self.status_label.setText(f"point_time 欄位轉換失敗: {e}")
            return

        filter_mode = self.filter_mode_combo.currentIndex()
        base_date = self.date_edit.date().toPyDate() if filter_mode == 1 else None

        # 取得補滿筆數
        fill_num = self.fillnum_spin.value()

        results = []

        if filter_mode == 0:
            # 全算
            grouped = df.groupby(["GroupName", "ChartName"])
            print("\n[DEBUG] All unique (GroupName, ChartName) pairs:")
            for pair in grouped.groups.keys():
                print("  ", pair)
            for (gname, cname), subdf in grouped:
                print(f"[DEBUG] Now processing group: GroupName='{gname}', ChartName='{cname}' | subdf.shape={subdf.shape}")
                characteristic = subdf["characteristic"].dropna().unique()
                if len(characteristic) != 1:
                    self.status_label.setText(f"Group: {gname}-{cname} 的 characteristic 不唯一或缺失")
                    continue
                group_stats = subdf.groupby("matching_group")["point_val"].agg(['mean', 'std', 'count']).reset_index()
                n_groups = len(group_stats)
                if n_groups == 2:
                    self._analyze_two_groups(group_stats, gname, cname, characteristic[0], results, subdf)
                else:
                    self._analyze_multiple_groups(subdf, group_stats, gname, cname, characteristic[0], results)
            self._create_boxplots(grouped)
        elif filter_mode == 1:
            # 指定日期模式
            grouped = df.groupby(["GroupName", "ChartName"])
            print("\n[DEBUG] All unique (GroupName, ChartName) pairs:")
            for pair in grouped.groups.keys():
                print("  ", pair)
            sigma_df_all = []  # 收集所有半年資料
            mean_df_all = []   # 收集所有一個月(補到5筆)資料
            for (gname, cname), subdf in grouped:
                print(f"[DEBUG] Now processing group: GroupName='{gname}', ChartName='{cname}' | subdf.shape={subdf.shape}")
                characteristic = subdf["characteristic"].dropna().unique()
                if len(characteristic) != 1:
                    self.status_label.setText(f"Group: {gname}-{cname} 的 characteristic 不唯一或缺失")
                    continue
                mean_end = pd.Timestamp(base_date)
                sigma_end = pd.Timestamp(base_date)
                mean_start = mean_end - pd.DateOffset(months=1)
                sigma_start = sigma_end - pd.DateOffset(months=6)
                # 先抓初始區間
                mean_df = subdf[(subdf["point_time"] > mean_start) & (subdf["point_time"] <= mean_end)].copy()
                sigma_df = subdf[(subdf["point_time"] > sigma_start) & (subdf["point_time"] <= sigma_end)].copy()
                # 針對每個 matching_group 補足 mean_df（只在不足時才補）
                min_time = subdf["point_time"].min()
                for mg in subdf["matching_group"].unique():
                    mg_mean = mean_df[mean_df["matching_group"] == mg]
                    if len(mg_mean) < fill_num:
                        all_mg = subdf[subdf["matching_group"] == mg].sort_values("point_time")
                        cur_start = mean_start
                        while len(mg_mean) < fill_num and cur_start > min_time:
                            cur_start = cur_start - pd.Timedelta(days=7)
                            mg_mean = all_mg[(all_mg["point_time"] > cur_start) & (all_mg["point_time"] <= mean_end)]
                        # 合併補足
                        mean_df = pd.concat([mean_df, mg_mean]).drop_duplicates()
                # sigma_df同理（只在不足時才補）
                for mg in subdf["matching_group"].unique():
                    mg_sigma = sigma_df[sigma_df["matching_group"] == mg]
                    if len(mg_sigma) < fill_num:
                        all_mg = subdf[subdf["matching_group"] == mg].sort_values("point_time")
                        cur_start = sigma_start
                        while len(mg_sigma) < fill_num and cur_start > min_time:
                            cur_start = cur_start - pd.Timedelta(days=14)
                            mg_sigma = all_mg[(all_mg["point_time"] > cur_start) & (all_mg["point_time"] <= sigma_end)]
                        sigma_df = pd.concat([sigma_df, mg_sigma]).drop_duplicates()
                mean_df_all.append(mean_df.assign(GroupName=gname, ChartName=cname))
                sigma_df_all.append(sigma_df.assign(GroupName=gname, ChartName=cname))
                mean_stats = mean_df.groupby("matching_group")["point_val"].agg(['mean', 'count']).reset_index()
                sigma_stats = sigma_df.groupby("matching_group")["point_val"].agg(['std']).reset_index()
                group_stats = pd.merge(mean_stats, sigma_stats, on="matching_group", how="outer")
                group_stats = group_stats.fillna({"mean": 0, "std": 0, "count": 0})
                n_groups = len(group_stats)
                if n_groups == 2:
                    # 對於指定日期模式，使用合併的數據進行統計檢定
                    combined_data = pd.concat([mean_df, sigma_df]).drop_duplicates()
                    self._analyze_two_groups(group_stats, gname, cname, characteristic[0], results, combined_data)
                else:
                    self._analyze_multiple_groups_time(mean_df, sigma_df, group_stats, gname, cname, characteristic[0], results)
            if mean_df_all:
                mean_df_concat = pd.concat(mean_df_all, ignore_index=True)
                mean_grouped = mean_df_concat.groupby(["GroupName", "ChartName"])
                self._create_boxplots(mean_grouped)
            else:
                self._create_boxplots(grouped)
        elif filter_mode == 2:
            # 最新進點模式
            grouped = df.groupby(["GroupName", "ChartName"])
            sigma_df_all = []
            mean_df_all = []
            for (gname, cname), subdf in grouped:
                characteristic = subdf["characteristic"].dropna().unique()
                if len(characteristic) != 1:
                    self.status_label.setText(f"Group: {gname}-{cname} 的 characteristic 不唯一或缺失")
                    continue
                latest_time = subdf["point_time"].max()
                mean_end = latest_time
                sigma_end = latest_time
                mean_start = mean_end - pd.DateOffset(months=1)
                sigma_start = sigma_end - pd.DateOffset(months=6)
                mean_df = subdf[(subdf["point_time"] > mean_start) & (subdf["point_time"] <= mean_end)].copy()
                sigma_df = subdf[(subdf["point_time"] > sigma_start) & (subdf["point_time"] <= sigma_end)].copy()
                min_time = subdf["point_time"].min()
                for mg in subdf["matching_group"].unique():
                    mg_mean = mean_df[mean_df["matching_group"] == mg]
                    if len(mg_mean) < fill_num:
                        all_mg = subdf[subdf["matching_group"] == mg].sort_values("point_time")
                        cur_start = mean_start
                        while len(mg_mean) < fill_num and cur_start > min_time:
                            cur_start = cur_start - pd.Timedelta(days=7)
                            mg_mean = all_mg[(all_mg["point_time"] > cur_start) & (all_mg["point_time"] <= mean_end)]
                        mean_df = pd.concat([mean_df, mg_mean]).drop_duplicates()
                for mg in subdf["matching_group"].unique():
                    mg_sigma = sigma_df[sigma_df["matching_group"] == mg]
                    if len(mg_sigma) < fill_num:
                        all_mg = subdf[subdf["matching_group"] == mg].sort_values("point_time")
                        cur_start = sigma_start
                        while len(mg_sigma) < fill_num and cur_start > min_time:
                            cur_start = cur_start - pd.Timedelta(days=14)
                            mg_sigma = all_mg[(all_mg["point_time"] > cur_start) & (all_mg["point_time"] <= sigma_end)]
                        sigma_df = pd.concat([sigma_df, mg_sigma]).drop_duplicates()
                mean_df_all.append(mean_df.assign(GroupName=gname, ChartName=cname))
                sigma_df_all.append(sigma_df.assign(GroupName=gname, ChartName=cname))
                mean_stats = mean_df.groupby("matching_group")["point_val"].agg(['mean', 'count']).reset_index()
                sigma_stats = sigma_df.groupby("matching_group")["point_val"].agg(['std']).reset_index()
                group_stats = pd.merge(mean_stats, sigma_stats, on="matching_group", how="outer")
                group_stats = group_stats.fillna({"mean": 0, "std": 0, "count": 0})
                n_groups = len(group_stats)
                if n_groups == 2:
                    # 對於最新進點模式，使用合併的數據進行統計檢定
                    combined_data = pd.concat([mean_df, sigma_df]).drop_duplicates()
                    self._analyze_two_groups(group_stats, gname, cname, characteristic[0], results, combined_data)
                else:
                    self._analyze_multiple_groups_time(mean_df, sigma_df, group_stats, gname, cname, characteristic[0], results)
            if mean_df_all:
                mean_df_concat = pd.concat(mean_df_all, ignore_index=True)
                mean_grouped = mean_df_concat.groupby(["GroupName", "ChartName"])
                self._create_boxplots(mean_grouped)
            else:
                self._create_boxplots(grouped)

        self._display_results(results)

    def _analyze_multiple_groups_time(self, mean_df, sigma_df, group_stats, gname, cname, characteristic, results):
        """
        多組分析（mean/std/count 來自一個月 window，median(sigma) 來自半年 window）
        - mean, std, count: 來自 mean_df（一個月 window，補到5筆）
        - median_sigma: 來自 sigma_df（半年 window，補到5筆）
        """
        # 只納入樣本數 >= 5 的 group 計算 median
        valid_mean_df = mean_df.groupby("matching_group").filter(lambda x: len(x) >= 5)
        sigma_by_group = sigma_df.groupby("matching_group")["point_val"].std()
        valid_groups = group_stats[group_stats['count'] >= 5]['matching_group']
        valid_sigma = sigma_by_group[valid_groups] if not valid_groups.empty else pd.Series(dtype=float)
        # 防呆：如果有效 group 只有一個，全部標記資料不足
        if len(valid_groups) <= 1:
            for i, row in group_stats.iterrows():
                group = row["matching_group"]
                mean = row["mean"]
                std = row["std"]
                n = row["count"]
                results.append([
                    gname, cname, group, "group_all",
                    '資料不足', '資料不足', 
                    self.get_k_value(n), mean, std, 
                    '-', '-', n
                ])
            return
        mean_median = valid_mean_df["point_val"].median() if not valid_mean_df.empty else 0
        median_sigma = valid_sigma.median() if not valid_sigma.empty else 0
        for i, row in group_stats.iterrows():
            group = row["matching_group"]
            mean = row["mean"]
            std = row["std"]  # 這是來自 mean_df（一個月 window）
            n = row["count"]
            if n < 5:
                results.append([
                    gname, cname, group, "group_all",
                    '資料不足', '資料不足', 
                    self.get_k_value(n), mean, std, 
                    mean_median, median_sigma, n
                ])
                continue
            if median_sigma > 0:
                if characteristic == 'up':
                    mean_index = (mean_median - mean) / median_sigma
                elif characteristic == 'down':
                    mean_index = (mean - mean_median) / median_sigma
                else:
                    mean_index = abs(mean - mean_median) / median_sigma
                sigma_index = std / median_sigma
            else:
                # 分母為零時，判斷所有 mean 是否相等
                all_means = group_stats['mean'].tolist() if not group_stats.empty else [mean]
                if len(set([round(m, 8) for m in all_means])) == 1:
                    mean_index = 0
                    sigma_index = 0
                else:
                    mean_index = float('inf')
                    sigma_index = float('inf')
            K = self.get_k_value(n)
            if K == "不比較":
                results.append([
                    gname, cname, group, "group_all",
                    '資料不足', '資料不足', 
                    '不比較', round(mean, 2), round(std, 2), 
                    round(mean_median, 2), round(median_sigma, 2), n
                ])
            else:
                results.append([
                    gname, cname, group, "group_all",
                    round(mean_index, 2), round(sigma_index, 2), 
                    round(K, 2), round(mean, 2), round(std, 2), 
                    round(mean_median, 2), round(median_sigma, 2), n
                ])

    def _analyze_two_groups(self, group_stats, gname, cname, characteristic, results, raw_data=None):
        """分析兩台設備的匹配情況"""
        row1 = group_stats.iloc[0]
        row2 = group_stats.iloc[1]

        group1 = row1["matching_group"]
        group2 = row2["matching_group"]
        mean1, std1, n1 = row1["mean"], row1["std"], row1["count"]
        mean2, std2, n2 = row2["mean"], row2["std"], row2["count"]

        min_sigma = min(std1, std2)

        # 檢查是否使用統計檢定
        use_stats = self.use_statistical_test_checkbox.isChecked()
        
        if use_stats and raw_data is not None and SCIPY_AVAILABLE:
            # 準備統計檢定的數據
            group1_data = raw_data[raw_data["matching_group"] == group1]["point_val"].tolist()
            group2_data = raw_data[raw_data["matching_group"] == group2]["point_val"].tolist()
            
            if len(group1_data) >= 2 and len(group2_data) >= 2:
                # 執行統計檢定
                method_map = {
                    0: "unpaired",
                    1: "paired"
                }
                method = method_map.get(self.stats_method_combo.currentIndex(), "unpaired")
                alpha = self.alpha_level_spin.value()
                
                data_groups = {group1: group1_data, group2: group2_data}
                stats_result = self.perform_statistical_test(data_groups, method, alpha)
                
                # 根據統計檢定結果決定 mean matching 狀態
                if "error" in stats_result:
                    mean_status_1 = f"統計檢定失敗: {stats_result['error']}"
                    mean_status_2 = f"統計檢定失敗: {stats_result['error']}"
                else:
                    p_val = stats_result.get("p_value", 1.0)
                    if stats_result.get("significant", False):
                        mean_status_1 = f"顯著差異 (p={p_val:.4f})"
                        mean_status_2 = f"顯著差異 (p={p_val:.4f})"
                    else:
                        mean_status_1 = f"無顯著差異 (p={p_val:.4f})"
                        mean_status_2 = f"無顯著差異 (p={p_val:.4f})"
            else:
                mean_status_1 = "統計檢定-數據不足"
                mean_status_2 = "統計檢定-數據不足"
        else:
            # 使用原有的 Mean Index 計算
            if min_sigma > 0:
                mean_index_1 = abs(mean1 - mean2) / min_sigma
                mean_index_2 = abs(mean2 - mean1) / min_sigma
            else:
                all_means = [mean1, mean2]
                if len(set([round(m, 8) for m in all_means])) == 1:
                    mean_index_1 = 0
                    mean_index_2 = 0
                else:
                    mean_index_1 = float('inf')
                    mean_index_2 = float('inf')
            
            mean_status_1 = round(mean_index_1, 2)
            mean_status_2 = round(mean_index_2, 2)

        # 計算 sigma index (保持原邏輯)
        if min_sigma > 0:
            sigma_index_1 = std1 / min_sigma
            sigma_index_2 = std2 / min_sigma
        else:
            sigma_index_1 = 0 if std1 == std2 else float('inf')
            sigma_index_2 = 0 if std1 == std2 else float('inf')

        # 統一格式：第4欄都用 'group_all'，與多群分析一致
        # mean_median, sigma_median 欄位（兩組時用 mean2, min_sigma 或 mean1, min_sigma）
        # 這裡用 mean2, min_sigma for group1, mean1, min_sigma for group2

        if n1 < 5 or n2 < 5:
            results.append([
                gname, cname, group1, 'group_all',
                '資料不足', '資料不足',
                self.get_k_value(n1), mean1, std1,
                mean2, min_sigma, n1
            ])
            results.append([
                gname, cname, group2, 'group_all',
                '資料不足', '資料不足',
                self.get_k_value(n2), mean2, std2,
                mean1, min_sigma, n2
            ])
            return

        k1 = self.get_k_value(n1)
        k2 = self.get_k_value(n2)

        if k1 == "不比較":
            results.append([
                gname, cname, group1, 'group_all',
                '資料不足', '資料不足',
                '不比較', round(mean1, 2), round(std1, 2),
                round(mean2, 2), round(min_sigma, 2), n1
            ])
        else:
            results.append([
                gname, cname, group1, 'group_all',
                mean_status_1, round(sigma_index_1, 2),
                round(k1, 2), round(mean1, 2), round(std1, 2),
                round(mean2, 2), round(min_sigma, 2), n1
            ])

        if k2 == "不比較":
            results.append([
                gname, cname, group2, 'group_all',
                '資料不足', '資料不足',
                '不比較', round(mean2, 2), round(std2, 2),
                round(mean1, 2), round(min_sigma, 2), n2
            ])
        else:
            results.append([
                gname, cname, group2, 'group_all',
                mean_status_2, round(sigma_index_2, 2),
                round(k2, 2), round(mean2, 2), round(std2, 2),
                round(mean1, 2), round(min_sigma, 2), n2
            ])

    def _analyze_multiple_groups(self, subdf, group_stats, gname, cname, characteristic, results):
        """分析多台設備的匹配情況 (mean matching index 分母都用 median_sigma)"""
        # 只納入樣本數 >= 5 的 group 計算 median
        valid_stats = group_stats[group_stats['count'] >= 5]
        if valid_stats.shape[0] <= 1:
            # 只有一個有效群組，全部標記資料不足
            for i, row in group_stats.iterrows():
                group = row["matching_group"]
                mean = row["mean"]
                std = row["std"]
                n = row["count"]
                results.append([
                    gname, cname, group, "group_all",
                    '資料不足', '資料不足', 
                    self.get_k_value(n), mean, std, 
                    '-', '-', n
                ])
            return

        mean_median = valid_stats['mean'].median() if not valid_stats.empty else 0
        median_sigma = valid_stats['std'].median() if not valid_stats.empty else 0

        # 檢查是否使用統計檢定
        use_stats = self.use_statistical_test_checkbox.isChecked()
        stats_result = None
        
        if use_stats and SCIPY_AVAILABLE and len(valid_stats) >= 3:
            # 準備 ANOVA 的數據
            data_groups = {}
            for _, row in valid_stats.iterrows():
                group_name = row["matching_group"]
                group_data = subdf[subdf["matching_group"] == group_name]["point_val"].tolist()
                if len(group_data) >= 2:
                    data_groups[group_name] = group_data
            
            if len(data_groups) >= 3:
                alpha = self.alpha_level_spin.value()
                stats_result = self.perform_statistical_test(data_groups, "unpaired", alpha)

        for i, row in group_stats.iterrows():
            group = row["matching_group"]
            mean = row["mean"]
            std = row["std"]
            n = row["count"]

            # 計算 mean matching index（考慮方向性）
            if n < 5:  # 樣本數不足5個，不進行比較
                results.append([
                    gname, cname, group, "group_all",
                    '資料不足', '資料不足', 
                    self.get_k_value(n), mean, std, 
                    mean_median, median_sigma, n
                ])
                continue

            # 決定 mean matching 狀態
            if use_stats and stats_result and "error" not in stats_result:
                # 使用統計檢定結果
                if stats_result.get("significant", False):
                    # ANOVA 顯著，檢查事後檢定結果
                    mean_status = "ANOVA顯著"
                    if stats_result.get("post_hoc") and "significant_pairs" in stats_result["post_hoc"]:
                        # 檢查這個群組是否在顯著配對中
                        significant_pairs = stats_result["post_hoc"]["significant_pairs"]
                        group_in_significant = any(
                            pair["group1"] == group or pair["group2"] == group 
                            for pair in significant_pairs
                        )
                        if group_in_significant:
                            # 找出與此群組有顯著差異的其他群組
                            different_groups = []
                            for pair in significant_pairs:
                                if pair["group1"] == group:
                                    different_groups.append(pair["group2"])
                                elif pair["group2"] == group:
                                    different_groups.append(pair["group1"])
                            if different_groups:
                                mean_status = f"與{','.join(different_groups)}有顯著差異"
                            else:
                                mean_status = "參與顯著差異"
                        else:
                            mean_status = "無顯著差異"
                    else:
                        mean_status = f"ANOVA顯著(p={stats_result.get('p_value', 0):.4f})"
                else:
                    mean_status = f"無顯著差異(p={stats_result.get('p_value', 1):.4f})"
            else:
                # 使用原有的 Mean Index 計算
                if median_sigma > 0:
                    if characteristic == 'up':
                        mean_index = (mean_median - mean) / median_sigma
                    elif characteristic == 'down':
                        mean_index = (mean - mean_median) / median_sigma
                    else:
                        mean_index = abs(mean - mean_median) / median_sigma
                else:
                    # 分母為零時，判斷所有 mean 是否相等
                    all_means = group_stats['mean'].tolist() if not group_stats.empty else [mean]
                    if len(set([round(m, 8) for m in all_means])) == 1:
                        mean_index = 0
                    else:
                        mean_index = float('inf')
                
                mean_status = round(mean_index, 2)

            # 計算 sigma index (保持原邏輯)
            if median_sigma > 0:
                sigma_index = std / median_sigma
            else:
                all_means = group_stats['mean'].tolist() if not group_stats.empty else [mean]
                if len(set([round(m, 8) for m in all_means])) == 1:
                    sigma_index = 0
                else:
                    sigma_index = float('inf')

            K = self.get_k_value(n)

            # 檢查 K 值是否為字串 "不比較"
            if K == "不比較":
                # 樣本數不足，使用 "資料不足" 標記
                results.append([
                    gname, cname, group, "group_all",
                    '資料不足', '資料不足', 
                    '不比較', round(mean, 2), round(std, 2), 
                    round(mean_median, 2), round(median_sigma, 2), n
                ])
            else:
                # 正常比較情況
                # 無論是否匹配都添加結果，保證所有比較都出現在報表中
                results.append([
                    gname, cname, group, "group_all",
                    mean_status, round(sigma_index, 2), 
                    round(K, 2), round(mean, 2), round(std, 2), 
                    round(mean_median, 2), round(median_sigma, 2), n
                ])

    def _display_results(self, results):
        """以新格式顯示分析結果，並在表格中添加按鈕以查看詳情。"""
        # 儲存報告數據以供彈出視窗使用
        self.report_data = {}
        
        # 遍歷結果，整理報表資料
        for row in results:
            gname, cname = row[0], row[1]
            key = f"{gname}_{cname}"
            
            if key not in self.report_data:
                self.report_data[key] = {
                    "GroupName": gname,
                    "ChartName": cname,
                    "groups": {}
                }
            
            group1, group2 = row[2], row[3]
            mean_index = row[4]
            sigma_index = row[5]
            
            if len(row) >= 12:
                k_value, mean, sigma, mean_median, sigma_median, n = row[6:12]
            else:
                k_value, mean, sigma, mean_median, sigma_median, n = [""] * 5 + [row[6] if len(row) > 6 else ""]
            
            if group2 == "group_all":
                self.report_data[key]["groups"][group1] = {
                    "mean_matching_index": mean_index,
                    "sigma_matching_index": sigma_index,
                    "K": k_value,
                    "mean": mean,
                    "sigma": sigma,
                    "mean_median": mean_median,
                    "sigma_median": sigma_median,
                    "samplesize": n
                }
            else:
                if group1 not in self.report_data[key]["groups"]:
                    self.report_data[key]["groups"][group1] = {}
                self.report_data[key]["groups"][group1][group2] = {
                    "mean_matching_index": mean_index,
                    "sigma_matching_index": sigma_index,
                    "K": k_value,
                    "mean": mean,
                    "sigma": sigma,
                    "mean_median": mean_median,
                    "sigma_median": sigma_median,
                    "samplesize": n
                }

        all_table_rows = []
        abnormal_ui_rows = []
        
        for key, data in self.report_data.items():
            gname = data["GroupName"]
            cname = data["ChartName"]
            
            for group_id, stats in data["groups"].items():
                mean_index = stats.get("mean_matching_index", "")
                sigma_index = stats.get("sigma_matching_index", "")
                k_value = stats.get("K", "")
                
                is_abnormal = False
                is_data_insufficient = mean_index == '資料不足' or sigma_index == '資料不足' or k_value == '不比較'
                abnormal_type = ""
                
                # 檢查是否為統計檢定的顯著差異
                is_statistical_significant = False
                if isinstance(mean_index, str) and ("顯著差異" in mean_index or "ANOVA顯著" in mean_index):
                    # 排除「無顯著差異」的情況
                    if "無顯著差異" not in mean_index:
                        is_statistical_significant = True
                    
                if not is_data_insufficient:
                    if is_statistical_significant:
                        # 統計檢定顯著
                        is_abnormal = True
                        abnormal_type = "統計檢定顯著"
                    else:
                        try:
                            # 只有勾選核取方塊才啟用門檻判斷，否則用預設值
                            mean_threshold = self.mean_index_threshold_spin.value() if self.mean_index_checkbox.isChecked() else 1.0
                            sigma_threshold = self.sigma_index_threshold_spin.value() if self.sigma_index_checkbox.isChecked() else float(k_value) if k_value not in [None, '', '不比較'] else 2.0
                            mean_abn = float(mean_index) >= mean_threshold
                            sigma_abn = float(sigma_index) >= sigma_threshold
                            if mean_abn or sigma_abn:
                                is_abnormal = True
                                if mean_abn and sigma_abn:
                                    abnormal_type = "Mean, Sigma"
                                elif mean_abn:
                                    abnormal_type = "Mean"
                                elif sigma_abn:
                                    abnormal_type = "Sigma"
                        except (ValueError, TypeError):
                            pass
                else:
                    abnormal_type = ""
                
                # 樣本數 n 強制轉為 int 顯示
                samplesize_val = stats.get("samplesize", "")
                try:
                    if samplesize_val != '' and samplesize_val is not None:
                        samplesize_val = int(float(samplesize_val))
                except Exception:
                    pass
                row_data = [
                    gname, cname, group_id,
                    stats.get("mean_matching_index", ""), stats.get("sigma_matching_index", ""),
                    stats.get("K", ""), stats.get("mean", ""), stats.get("sigma", ""),
                    stats.get("mean_median", ""), stats.get("sigma_median", ""),
                    samplesize_val
                ]
                
                all_row_data = [is_abnormal, abnormal_type] + row_data
                all_table_rows.append(all_row_data)
                
                if is_abnormal or is_data_insufficient:
                    abnormal_ui_rows.append({
                        "key": (gname, cname),
                        "group_id": group_id,
                        "data": [abnormal_type] + row_data
                    })
        

        # 填充表格 (只顯示異常項目)
        self.result_table.setColumnCount(13)
        self.result_table.setHorizontalHeaderLabels([
            "View Details", "Abnormal Type", "Group Name", "Chart Name", "Matching Group", "Mean Index", "Sigma Index",
            "K", "Mean", "Sigma", "Mean Median", "Sigma Median", "Sample Size"
        ])
        self.result_table.setRowCount(len(abnormal_ui_rows))


        for i, item_info in enumerate(abnormal_ui_rows):
            # 使用眼睛 icon 按鈕
            view_button = QtWidgets.QPushButton()
            eye_icon = self.style().standardIcon(QtWidgets.QStyle.StandardPixmap.SP_DesktopIcon)  # fallback 預設 icon
            # 嘗試用 PyQt6 內建的 eye icon，如果有
            try:
                eye_icon = self.style().standardIcon(QtWidgets.QStyle.StandardPixmap.SP_FileDialogContentsView)
            except Exception:
                pass
            view_button.setIcon(eye_icon)
            view_button.setToolTip("檢視詳細資訊")
            view_button.setFixedWidth(36)
            view_button.setFixedHeight(36)
            view_button.setIconSize(QtCore.QSize(22, 22))
            view_button.setStyleSheet("QPushButton { border: none; background: transparent; } QPushButton:hover { background: #e0e7ef; }")
            # 置中顯示
            cell_widget = QtWidgets.QWidget()
            layout = QtWidgets.QHBoxLayout(cell_widget)
            layout.setContentsMargins(0, 0, 0, 0)
            layout.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            layout.addWidget(view_button)
            view_button.clicked.connect(
                lambda checked, key=item_info["key"], gid=item_info["group_id"]: self._show_details_dialog(key, gid)
            )
            self.result_table.setCellWidget(i, 0, cell_widget)

            # 填充其他數據（多一欄異常類型）
            row_data = item_info["data"]
            for j, val in enumerate(row_data):
                # --- 格式化數值欄位為兩位小數 ---
                if j in [4,5,6,7,8,9,10]:  # Mean Index, Sigma Index, K, Mean, Sigma, Mean Median, Sigma Median
                    try:
                        if val != '資料不足' and val != '不比較' and val != '' and val is not None:
                            val = float(val)
                            val = f"{val:.2f}"
                    except Exception:
                        pass
                item = QtWidgets.QTableWidgetItem(str(val))
                item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
                # 標記異常值
                is_abnormal = False
                try:
                    # 檢查是否為統計檢定顯著差異
                    if j == 4 and isinstance(val, str) and ("顯著差異" in val or "ANOVA顯著" in val):
                        # 排除「無顯著差異」的情況
                        if "無顯著差異" not in val:
                            is_abnormal = True
                    else:
                        # 檢查傳統的 Mean Index 和 Sigma Index 門檻
                        mean_idx_val = float(row_data[4])
                        sigma_idx_val = float(row_data[5])
                        k_val = float(row_data[6])
                        if (j == 4 and mean_idx_val >= 1) or (j == 5 and sigma_idx_val >= k_val):
                            is_abnormal = True
                except (ValueError, TypeError):
                    pass
                # 只標記異常欄位為紅字，不設底色，避免 QSS 衝突
                if is_abnormal:
                    item.setForeground(QtGui.QColor("#D32F2F"))
                self.result_table.setItem(i, j + 1, item)

        self.result_table.resizeColumnsToContents()
        self.result_table.horizontalHeader().setStretchLastSection(True)

        # 匯出全部結果到 Excel 檔案
        if all_table_rows and hasattr(self, 'file_path_entry') and self.file_path_entry.text():
            self._export_to_excel(all_table_rows, self.file_path_entry.text())
        else:
            self.status_label.setText(f"分析完成，共發現 {len(abnormal_ui_rows)} 個需注意項目。")
            
        if len(abnormal_ui_rows) > 0:
            self.status_label.setText(f"分析完成，共發現 {len(abnormal_ui_rows)} 個需注意項目（總共 {len(all_table_rows)} 項）。")
        else:
            self.status_label.setText(f"分析完成，未發現需注意項目（總共 {len(all_table_rows)} 項）。")

    def _show_details_dialog(self, chart_key, group_id):
        """彈出一個視窗，顯示詳細資訊和圖表，上方為數據，下方為圖表。"""
        try:
            from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
        except ImportError:
            QtWidgets.QMessageBox.warning(self, "缺少套件", "顯示圖表需要 Matplotlib。")
            return

        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle(f"詳細資訊: {chart_key[0]} - {chart_key[1]} | 組別: {group_id}")
        dialog.setMinimumSize(1400, 450) # 調整視窗大小以適應新佈局 (高度減少)

        main_layout = QtWidgets.QVBoxLayout(dialog)
        main_layout.setSpacing(10)

        # --- 上方：數據表格 (水平排列) ---
        try:
            stats = self.report_data[f"{chart_key[0]}_{chart_key[1]}"]["groups"][group_id]
        except KeyError:
            QtWidgets.QMessageBox.critical(self, "錯誤", "找不到此項目的詳細數據。")
            return


        info_group = QtWidgets.QGroupBox("Analysis Data")
        info_v_layout = QtWidgets.QVBoxLayout(info_group)

        # 取得異常類型
        # 這裡需重算異常類型，與 UI/Excel 一致
        mean_index = stats.get("mean_matching_index", "")
        sigma_index = stats.get("sigma_matching_index", "")
        k_value = stats.get("K", "")
        abnormal_type = ""
        is_data_insufficient = mean_index == '資料不足' or sigma_index == '資料不足' or k_value == '不比較'
        
        # 檢查是否為統計檢定的顯著差異
        is_statistical_significant = False
        if isinstance(mean_index, str) and ("顯著差異" in mean_index or "ANOVA顯著" in mean_index):
            # 排除「無顯著差異」的情況
            if "無顯著差異" not in mean_index:
                is_statistical_significant = True
            
        if not is_data_insufficient:
            if is_statistical_significant:
                # 統計檢定顯著
                abnormal_type = "統計檢定顯著"
            else:
                try:
                    mean_threshold = self.mean_index_threshold_spin.value() if self.mean_index_checkbox.isChecked() else 1.0
                    sigma_threshold = self.sigma_index_threshold_spin.value() if self.sigma_index_checkbox.isChecked() else float(k_value) if k_value not in [None, '', '不比較'] else 2.0
                    mean_abn = float(mean_index) >= mean_threshold
                    sigma_abn = float(sigma_index) >= sigma_threshold
                    if mean_abn and sigma_abn:
                        abnormal_type = "Mean, Sigma"
                    elif mean_abn:
                        abnormal_type = "Mean"
                    elif sigma_abn:
                        abnormal_type = "Sigma"
                except (ValueError, TypeError):
                    pass

        # 新增異常類型欄位
        headers = [
            "Abnormal Type", "Group Name", "Chart Name", "Matching Group", "Mean Index", "Sigma Index",
            "K", "Mean", "Sigma", "Mean Median", "Sigma Median", "Sample Size"
        ]
        gname, cname = chart_key
        # 樣本數 n 強制轉為 int 顯示
        samplesize_val = stats.get("samplesize", "")
        try:
            if samplesize_val != '' and samplesize_val is not None:
                samplesize_val = int(float(samplesize_val))
        except Exception:
            pass
        row_values = [
            abnormal_type,
            gname, cname, group_id,
            mean_index, sigma_index,
            stats.get("K", ""), stats.get("mean", ""), stats.get("sigma", ""),
            stats.get("mean_median", ""), stats.get("sigma_median", ""),
            samplesize_val
        ]

        info_table = QtWidgets.QTableWidget()
        info_table.setColumnCount(len(headers))
        info_table.setHorizontalHeaderLabels(headers)
        info_table.setRowCount(1)
        info_table.verticalHeader().setVisible(False)
        info_table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)

        for j, value in enumerate(row_values):
            if j in [4,5,6,7,8,9,10]:  # Mean Index, Sigma Index, K, Mean, Sigma, Mean Median, Sigma Median
                try:
                    if value != '資料不足' and value != '不比較' and value != '' and value is not None:
                        value = float(value)
                        value = f"{value:.2f}"
                except Exception:
                    pass
            item = QtWidgets.QTableWidgetItem(str(value))
            item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            info_table.setItem(0, j, item)

        info_table.resizeColumnsToContents()
        info_table.setFixedHeight(info_table.horizontalHeader().height() + info_table.rowHeight(0) + 5)
        info_v_layout.addWidget(info_table)
        main_layout.addWidget(info_group)

        # --- 下方：圖表區塊 ---
        charts_container_widget = QtWidgets.QWidget()
        charts_layout = QtWidgets.QHBoxLayout(charts_container_widget)

        if hasattr(self, 'chart_figures') and chart_key in self.chart_figures:
            figures = self.chart_figures[chart_key]
            
            if figures['scatter'] and figures['box']:
                # --- 解決圖表重複開啟變大問題 ---
                # 使用 pickle 進行深度複製，確保每次顯示都是全新的 Figure 物件
                scatter_fig_copy = pickle.loads(pickle.dumps(figures['scatter']))
                box_fig_copy = pickle.loads(pickle.dumps(figures['box']))

                scatter_canvas = FigureCanvas(scatter_fig_copy)
                box_canvas = FigureCanvas(box_fig_copy)
                # ------------------------------------
                
                scatter_canvas.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Expanding)
                box_canvas.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Expanding)

                charts_layout.addWidget(scatter_canvas)
                charts_layout.addWidget(box_canvas)
            else:
                charts_layout.addWidget(QtWidgets.QLabel("此項目的圖表因數據不足未生成。"))
        else:
            charts_layout.addWidget(QtWidgets.QLabel("找不到對應的圖表。"))
        
        main_layout.addWidget(charts_container_widget)

        # 設定佈局伸展因子，讓圖表區域佔用更多空間
        main_layout.setStretchFactor(info_group, 0) # 數據表格高度固定
        main_layout.setStretchFactor(charts_container_widget, 1) # 圖表區域填滿剩餘空間

        # --- 關閉按鈕 ---
        button_box = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.StandardButton.Close)
        button_box.rejected.connect(dialog.reject)
        main_layout.addWidget(button_box)

        dialog.exec()

    def _create_boxplots(self, grouped):
        """創建 SPC 圖和盒鬚圖，將 figure 物件保存在 self.chart_figures 中，不在 UI 上顯示。"""
        try:
            # 這些導入是必要的，因為 Matplotlib 在子線程或不同上下文中可能需要重新導入
            import matplotlib.pyplot as plt
            from matplotlib import cm
            import numpy as np
        except ImportError:
            print("[ERROR] Matplotlib is not installed.")
            return

        # 保存圖表與分組鍵的對應關係，用於後續的彈出視窗和 Excel 匯出
        self.chart_figures = {}
        
        # 為每個 (GroupName, ChartName) 組合創建圖表
        for (gname, cname), subdf in grouped:
            # 依 matching_group 字母順序排序
            unique_groups = sorted(subdf["matching_group"].unique(), key=lambda x: str(x))
            labels = [str(mg) for mg in unique_groups]

            # 檢查是否有數據可供繪圖
            if subdf.empty or not any(len(grp["point_val"]) > 0 for _, grp in subdf.groupby("matching_group")):
                print(f"[WARNING] Skipping chart creation for {gname} - {cname} due to empty data.")
                self.chart_figures[(gname, cname)] = {'scatter': None, 'box': None}
                continue

            # 依排序後 unique_groups 組裝 box_data，確保顏色/label/資料一致
            box_data = [subdf[subdf["matching_group"] == mg]["point_val"].values for mg in unique_groups]
            group_stats = subdf.groupby("matching_group")["point_val"].agg(['mean', 'std', 'count'])

            # 為不同的組設置顏色
            colors = cm.tab10(np.linspace(0, 1, len(unique_groups)))

            # 1. 創建 SPC 風格的圖表
            scatter_fig, scatter_ax = plt.subplots(figsize=(7, 4.5)) # 調整尺寸為較小的長方形
            
            # 計算整體統計量用於控制線
            all_values = subdf["point_val"].values
            # overall_mean = np.mean(all_values)
            # overall_std = np.std(all_values)
            

            # 為每個群組繪製數據點，按時間順序連線
            x_position = 0
            for i, mg in enumerate(unique_groups):
                group_data = subdf[subdf["matching_group"] == mg].sort_values("point_time")
                if not group_data.empty:
                    # 為每個群組創建連續的x位置
                    x_vals = np.arange(x_position, x_position + len(group_data))
                    y_vals = group_data["point_val"].values
                    
                    # 繪製數據點
                    scatter_ax.scatter(x_vals, y_vals, color=colors[i], alpha=0.8, s=40, label=f'{mg}', zorder=3)
                    
                    # 連接同組內的點
                    scatter_ax.plot(x_vals, y_vals, color=colors[i], alpha=0.5, linewidth=1, zorder=2)
                    
                    # 在群組間添加分隔線
                    if i < len(unique_groups) - 1:  # 不在最後一組後面加線
                        separator_x = x_position + len(group_data) - 0.5
                        scatter_ax.axvline(x=separator_x, color='gray', linestyle='-', alpha=0.3, zorder=1)
                    
                    x_position += len(group_data)
            
            # 設置圖表樣式
            scatter_ax.set_title(f"SPC Chart: {gname} - {cname}", fontsize=10)
            scatter_ax.set_xlabel("Sample Sequence (Grouped by Matching Group)")
            scatter_ax.set_ylabel("Point Value")
            scatter_ax.grid(True, linestyle='--', alpha=0.3, zorder=0)
            
            # 添加群組標籤在x軸上
            if unique_groups:
                group_positions = []
                x_pos = 0
                for mg in unique_groups:
                    group_size = len(subdf[subdf["matching_group"] == mg])
                    group_positions.append(x_pos + group_size/2 - 0.5)
                    x_pos += group_size
                
                # 設置x軸刻度和標籤
                scatter_ax.set_xticks(group_positions)
                scatter_ax.set_xticklabels(labels, rotation=0, ha='center')
                
                # 添加次要刻度顯示樣本序號
                scatter_ax.tick_params(axis='x', which='minor', bottom=True, top=False)
            
            # 調整圖例位置
            scatter_ax.legend(loc='upper left', bbox_to_anchor=(1.02, 1), fontsize='small')
            scatter_fig.tight_layout()

            # 2. 創建盒鬚圖
            box_fig, box_ax = plt.subplots(figsize=(7, 4.5)) # 調整尺寸為較小的長方形
            if box_data:
                bp = box_ax.boxplot(box_data, labels=labels, patch_artist=True, widths=0.6)
                for patch, color in zip(bp['boxes'], colors):
                    patch.set_facecolor(color)

                # legend 也照 unique_groups 順序
                legend_labels = [
                    f"{label}: μ={group_stats.loc[mg, 'mean']:.2f}, σ={group_stats.loc[mg, 'std']:.2f}, n={int(group_stats.loc[mg, 'count'])}"
                    for label, mg in zip(labels, unique_groups)
                ]
                box_ax.legend([bp["boxes"][i] for i in range(len(labels))], legend_labels, loc='upper left', bbox_to_anchor=(1.02, 1), fontsize='small')

            box_ax.set_title(f"Boxplot: {gname} - {cname}", fontsize=10)
            box_ax.set_xlabel("Matching Group")
            box_ax.set_ylabel("Point Value")
            box_ax.grid(True, linestyle='--', alpha=0.6)
            box_fig.subplots_adjust(right=0.7)
            box_fig.tight_layout()

            # 保存圖表與分組鍵的映射
            key = (gname, cname)
            self.chart_figures[key] = {'scatter': scatter_fig, 'box': box_fig}  # scatter實際上是SPC圖

            # 關鍵：關閉 figure 以釋放記憶體，因為我們已經將其保存在 self.chart_figures 中
            # FigureCanvas 會在需要時重新繪製它
            plt.close(scatter_fig)
            plt.close(box_fig)

    def _export_to_excel(self, all_results, source_path):
        """將分析結果匯出為 Excel 檔案，並在第一欄嵌入完整的盒鬚圖和散點圖。包含異常類型欄。"""
        try:
            # 檢查是否已安裝 openpyxl
            if openpyxl is None:
                QtWidgets.QMessageBox.warning(
                    self, "缺少套件", 
                    "請安裝 openpyxl 以匯出 Excel 檔案。\n可在終端執行: pip install openpyxl"
                )
                self.status_label.setText(f"分析完成。無法匯出 Excel：需要 openpyxl 套件。")
                return None

            # 嘗試導入所需的模組
            try:
                import matplotlib.pyplot as plt
                import numpy as np
                import io
                from PIL import Image
                import matplotlib.cm as cm
                from openpyxl.drawing.image import Image as XLImage
            except ImportError as e:
                QtWidgets.QMessageBox.warning(
                    self, "缺少套件", 
                    f"嵌入圖表需要額外套件: {str(e)}\n請安裝所需套件。"
                )
                print(f"[WARNING] 缺少嵌入圖表所需套件: {e}")
                return None

            # 新增異常類型欄位，all_results: [is_abnormal, abnormal_type, ...]
            columns = [
                "Need_matching", "AbnormalType", "GroupName", "ChartName", "matching_group", "mean_matching_index", 
                "sigma_matching_index", "K", "mean", "sigma", "mean_median", "sigma_median", "samplesize"
            ]
            df = pd.DataFrame(all_results, columns=columns)

            # 打印資料框資訊以確認結構
            print(f"DataFrame info: {df.shape}")
            print(f"DataFrame columns: {df.columns.tolist()}")
            print(f"First row: {df.iloc[0].tolist() if len(df) > 0 else 'No data'}")

            # 生成輸出檔案路徑（與輸入檔案相同目錄）
            dir_path = os.path.dirname(source_path)
            file_name = os.path.splitext(os.path.basename(source_path))[0]
            output_path = os.path.join(dir_path, f"{file_name}_matching_results.xlsx")

            # 創建臨時目錄用於保存圖片
            import tempfile
            temp_dir = tempfile.mkdtemp()
            print(f"[INFO] 創建臨時目錄: {temp_dir}")

            # 先在 DataFrame 前添加兩個空白欄位，分別用於SPC圖和盒鬚圖
            df.insert(0, "SPC_Chart", "")    # 第一欄：SPC圖
            df.insert(1, "BoxPlot", "")      # 第二欄：盒鬚圖

            # 創建 Excel 文件
            writer = pd.ExcelWriter(output_path, engine='openpyxl')
            df.to_excel(writer, sheet_name='Tool Matching Results', index=False)

            # 獲取工作表
            workbook = writer.book
            worksheet = writer.sheets['Tool Matching Results']

            # 設定標題列格式
            header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            header_fill = openpyxl.styles.PatternFill(start_color="344CB7", end_color="344CB7", fill_type="solid")
            header_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

            # 設置標題列格式
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 增加圖表欄寬度以容納圖片
            worksheet.column_dimensions['A'].width = 70  # 第一欄：SPC圖
            worksheet.column_dimensions['B'].width = 70  # 第二欄：盒鬚圖

            # 設定異常行的格式
            abnormal_fill = openpyxl.styles.PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

            # 定義圖表在 Excel 中顯示的尺寸 (單位：像素)
            img_display_width, img_display_height = 450, 250

            # 檢查是否有可用的圖表數據
            has_chart_figures = hasattr(self, 'chart_figures') and self.chart_figures
            if not has_chart_figures:
                print("[WARNING] 沒有可用的圖表數據，將使用簡單的狀態指示圖")

            # 從第二行開始遍歷（跳過標題行）
            for row_idx, row in enumerate(df.iterrows(), start=2):
                _, row_data = row

                # 檢查Need_matching欄位是否為True
                is_abnormal = row_data["Need_matching"]

                if is_abnormal:
                    # 將整行設為淺紅色
                    for cell in worksheet[row_idx]:
                        cell.fill = abnormal_fill

                # 創建並嵌入圖表到第一欄
                try:
                    # 獲取關鍵數據
                    group_name = str(row_data["GroupName"])
                    chart_name = str(row_data["ChartName"])
                    group_id = str(row_data["matching_group"])
                    mean_index = row_data["mean_matching_index"]
                    sigma_index = row_data["sigma_matching_index"]
                    k_value = row_data["K"]

                    # 檢查是否資料不足
                    is_data_insufficient = (mean_index == '資料不足' or sigma_index == '資料不足' or k_value == '不比較')

                    # 嘗試使用完整的SPC圖和盒鬚圖
                    chart_key = (group_name, chart_name)
                    if has_chart_figures and chart_key in self.chart_figures:
                        # 存在完整的分析圖表，使用實際的SPC圖和盒鬚圖
                        chart_data = self.chart_figures[chart_key]

                        # 1. 處理SPC圖 (放在第一欄)
                        try:
                            scatter_fig = chart_data['scatter']
                            temp_scatter_path = os.path.join(temp_dir, f"spc_{group_name}_{chart_name}_{row_idx}.png")
                            scatter_fig.savefig(temp_scatter_path, format='png', bbox_inches='tight', transparent=True, dpi=100)
                            try:
                                scatter_img = XLImage(temp_scatter_path)
                                scatter_img.width = img_display_width
                                scatter_img.height = img_display_height
                                scatter_position = f"A{row_idx}"
                                worksheet.add_image(scatter_img, scatter_position)
                                print(f"[INFO] 已添加SPC圖到單元格: {scatter_position}")
                            except Exception as img_e:
                                print(f"[ERROR] 添加SPC圖到 Excel 失敗: {img_e}")
                                worksheet.cell(row=row_idx, column=1).value = "SPC圖載入失敗"
                        except Exception as scatter_e:
                            print(f"[ERROR] 處理SPC圖時發生錯誤: {scatter_e}")
                            import traceback
                            traceback.print_exc()
                            worksheet.cell(row=row_idx, column=1).value = "SPC圖生成失敗"

                        # 2. 處理盒鬚圖 (放在第二欄)
                        try:
                            box_fig = chart_data['box']
                            temp_box_path = os.path.join(temp_dir, f"box_{group_name}_{chart_name}_{row_idx}.png")
                            box_fig.savefig(temp_box_path, format='png', bbox_inches='tight', transparent=True, dpi=100)
                            try:
                                box_img = XLImage(temp_box_path)
                                box_img.width = img_display_width
                                box_img.height = img_display_height
                                box_position = f"B{row_idx}"
                                worksheet.add_image(box_img, box_position)
                                print(f"[INFO] 已添加盒鬚圖到單元格: {box_position}")
                            except Exception as img_e:
                                print(f"[ERROR] 添加盒鬚圖到 Excel 失敗: {img_e}")
                                worksheet.cell(row=row_idx, column=2).value = "盒鬚圖載入失敗"
                        except Exception as box_e:
                            print(f"[ERROR] 處理盒鬚圖時發生錯誤: {box_e}")
                            import traceback
                            traceback.print_exc()
                            worksheet.cell(row=row_idx, column=2).value = "盒鬚圖生成失敗"

                    else:
                        # 沒有找到匹配的圖表，使用狀態指示器
                        print(f"[INFO] 未找到 {group_name}/{chart_name} 的分析圖表，使用狀態指示器")
                        fig, ax = plt.subplots(figsize=(6, 4), dpi=100)
                        title = f"{group_name}\n{chart_name}\n組別: {group_id}"
                        ax.set_title(title, fontsize=12)
                        if is_data_insufficient:
                            circle = plt.Circle((0.5, 0.5), 0.3, color='yellow', alpha=0.6, edgecolor='goldenrod', linewidth=2)
                            ax.add_patch(circle)
                            ax.text(0.5, 0.5, "資料不足", ha='center', va='center', fontsize=14, color='black')
                            status_text = "資料不足，無法進行分析"
                        elif is_abnormal:
                            circle = plt.Circle((0.5, 0.5), 0.3, color='red', alpha=0.6, edgecolor='darkred', linewidth=2)
                            ax.add_patch(circle)
                            ax.text(0.5, 0.5, "需要對齊", ha='center', va='center', fontsize=14, color='white', fontweight='bold')
                            status_text = f"均值差異指數: {mean_index}, 標準差差異指數: {sigma_index}, K值: {k_value}"
                        else:
                            circle = plt.Circle((0.5, 0.5), 0.3, color='green', alpha=0.6, edgecolor='darkgreen', linewidth=2)
                            ax.add_patch(circle)
                            ax.text(0.5, 0.5, "正常", ha='center', va='center', fontsize=14, color='white', fontweight='bold')
                            status_text = f"均值差異指數: {mean_index}, 標準差差異指數: {sigma_index}, K值: {k_value}"
                        ax.text(0.5, 0.2, status_text, ha='center', va='center', fontsize=10, 
                               bbox=dict(boxstyle='round,pad=0.5', facecolor='white', alpha=0.8))
                        ax.set_xticks([])
                        ax.set_yticks([])
                        ax.set_xlim(0, 1)
                        ax.set_ylim(0, 1)
                        ax.set_aspect('equal')
                        temp_img_path = os.path.join(temp_dir, f"status_chart_{row_idx}.png")
                        plt.savefig(temp_img_path, format='png', bbox_inches='tight', transparent=True, dpi=300)
                        plt.close(fig)
                        try:
                            # 使用 xlsxwriter 寫法 (insert_image) 取代 openpyxl 的 add_image
                            # 需先取得 xlsxwriter 的 worksheet 物件
                            # 但目前本程式是用 openpyxl，無法直接用 insert_image
                            # 所以這裡僅說明：如果你要用 insert_image，必須用 xlsxwriter 建立 writer
                            # 下面是 xlsxwriter 寫法範例：
                            # worksheet.insert_image(row_idx-1, 0, temp_img_path, {'x_scale': 1, 'y_scale': 1, 'x_offset': 0, 'y_offset': 0, 'object_position': 1})
                            # worksheet.insert_image(row_idx-1, 1, temp_img_path, {'x_scale': 1, 'y_scale': 1, 'x_offset': 0, 'y_offset': 0, 'object_position': 1})
                            # 但 openpyxl 不支援 insert_image，僅支援 add_image
                            # 若要完全改用 xlsxwriter，需重構整個 Excel 輸出流程。
                            # 這裡保留原本 openpyxl add_image 寫法，僅註明差異。
                            img1 = XLImage(temp_img_path)
                            img1.width = img_display_width
                            img1.height = img_display_height
                            cell_position_1 = f"A{row_idx}"
                            worksheet.add_image(img1, cell_position_1)
                            img2 = XLImage(temp_img_path)
                            img2.width = img_display_width
                            img2.height = img_display_height
                            cell_position_2 = f"B{row_idx}"
                            worksheet.add_image(img2, cell_position_2)
                            print(f"[INFO] 已添加狀態圖到單元格: {cell_position_1} 和 {cell_position_2}")
                        except Exception as img_e:
                            print(f"[ERROR] 添加圖片到 Excel 失敗: {img_e}")
                            worksheet.cell(row=row_idx, column=1).value = "圖片載入失敗"
                            worksheet.cell(row=row_idx, column=2).value = "圖片載入失敗"

                except Exception as img_e:
                    print(f"[ERROR] 在第 {row_idx} 行添加圖表時發生錯誤: {img_e}")
                    import traceback
                    traceback.print_exc()
                    worksheet.cell(row=row_idx, column=1).value = "圖片生成失敗"

            # 調整行高以適應圖表
            for i in range(2, worksheet.max_row + 1):
                worksheet.row_dimensions[i].height = 190

            # 調整其他列寬
            for col_idx, column in enumerate(worksheet.columns, start=1):
                if col_idx <= 2:  # 跳過圖表列 A 和 B，已手動設置寬度
                    continue
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 4)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # 儲存 Excel 檔案
            try:
                writer.close()
                print(f"[INFO] Excel 檔案已儲存到: {output_path}")
            except Exception as save_e:
                print(f"[ERROR] 儲存 Excel 檔案失敗: {save_e}")
                import traceback
                traceback.print_exc()
            finally:
                try:
                    import shutil
                    shutil.rmtree(temp_dir)
                    print(f"[INFO] 已清理臨時目錄: {temp_dir}")
                except Exception as e:
                    print(f"[WARNING] 無法清理臨時目錄: {temp_dir}, 錯誤: {e}")

            self.status_label.setText(f"分析完成。結果已匯出到: {output_path}")
            return output_path
        except Exception as e:
            self.status_label.setText(f"匯出 Excel 失敗: {e}")
            import traceback
            traceback.print_exc()
            return None